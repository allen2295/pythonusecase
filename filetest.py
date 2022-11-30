# 文件操作
'''
with open('./data/test.xlsx','rb') as f:
    for line in f.readlines():
        print(line.strip())
'''
'''
import os
import platform

print(os.name)

cudir = os.path.abspath('.')
print(cudir)

print(platform.processor())
'''
# test1
import os
'''
from openpyxl import load_workbook
wb = load_workbook('./data/test.xlsx')

ws = wb.active
sheet_name = wb.get_sheet_names()
print(sheet_name[0])
a_sheet = wb.get_sheet_by_name(sheet_name[0])
print(a_sheet.title)
C4 = a_sheet['C4']
print(f'({C4.column}, {C4.row}) is {C4.value}')
for row in a_sheet.rows:
    for cell in row:
        print(cell.value)
#print(wb.get)
'''

from openpyxl import load_workbook
wb = load_workbook('./data/test.xlsx')
#ws1 = wb.create_sheet('add1') #插到最后
#ws2 = wb.create_sheet('倒数第二个', -1)
print(wb.sheetnames)
sheet_name = wb.get_sheet_names()
ws2 = wb.get_sheet_by_name(sheet_name[2])
#ws2.title = '倒数第二个'
#ws2.sheet_properties.tabColor = "1072BA"
'''
for row in ws2.values:
    for x in row:
        print(x)
'''
for x in range(1, 11):
    for y in range(1, 11):
        tempcell = ws2.cell(row=x, column=y)
        print(tempcell.value)
        if ws2.cell(row=x, column=y).value == int(2):
           ws2.cell(row=x, column=y).value = 1111


wb.save('./data/test.xlsx')